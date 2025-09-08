from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import os
import tempfile
try:
    from ..models import db, MobileHome, StateRequirement, Property, User
    from ..utils import role_required
except ImportError:
    from models import db, MobileHome, StateRequirement, Property, User
    from utils import role_required

bp = Blueprint('excel_export', __name__, url_prefix='/api/excel')

def create_vin_recovery_worksheet(workbook, state_code, mobile_homes, state_req):
    """Create VIN/Title recovery worksheet for a specific state"""
    ws = workbook.create_sheet(title=f"{state_code} VIN Recovery")
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    ws.merge_cells('A1:P1')
    title_cell = ws['A1']
    title_cell.value = f"Mobile Home Park VIN/Title Recovery - {state_req.state_name if state_req else state_code}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = header_alignment
    title_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # State information section
    row = 3
    ws[f'A{row}'] = "State Agency:"
    ws[f'B{row}'] = state_req.agency_name if state_req else "N/A"
    row += 1
    ws[f'A{row}'] = "Processing Time:"
    ws[f'B{row}'] = state_req.processing_time if state_req else "N/A"
    row += 1
    ws[f'A{row}'] = "Website:"
    ws[f'B{row}'] = state_req.website_url if state_req else "N/A"
    row += 2
    
    # Headers for mobile home data
    headers = [
        "VIN Number", "Make", "Model", "Year", "Serial Number",
        "Current Owner", "Owner Address", "Owner Phone",
        "Previous Owner", "Title Status", "Lien Holder",
        "Lien Amount", "Acquisition Date", "Park Space",
        "Purchase Price", "Recovery Status"
    ]
    
    # Apply headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row += 1
    
    # Add mobile home data
    for mh in mobile_homes:
        data = [
            mh.vin_number or "",
            mh.make or "",
            mh.model or "",
            mh.year or "",
            mh.serial_number or "",
            mh.current_owner_name or "",
            mh.current_owner_address or "",
            mh.current_owner_phone or "",
            mh.previous_owner_name or "",
            mh.title_status or "",
            mh.lien_holder or "",
            mh.lien_amount or "",
            mh.acquisition_date.strftime('%m/%d/%Y') if mh.acquisition_date else "",
            mh.park_location_space or "",
            mh.purchase_price or "",
            mh.recovery_status or ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            if col in [12, 15]:  # Money columns
                cell.number_format = '$#,##0.00'
        row += 1
    
    # Add required documents section
    if state_req and state_req.required_documents:
        row += 2
        ws[f'A{row}'] = "Required Documents:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        required_docs = json.loads(state_req.required_documents)
        for doc in required_docs:
            ws[f'B{row}'] = f"• {doc}"
            row += 1
    
    # Add fees section
    if state_req and state_req.fees:
        row += 1
        ws[f'A{row}'] = "Fees:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        fees = json.loads(state_req.fees)
        for fee_type, amount in fees.items():
            ws[f'B{row}'] = f"{fee_type}: ${amount}"
            row += 1
    
    # Special requirements
    if state_req and state_req.special_requirements:
        row += 1
        ws[f'A{row}'] = "Special Requirements:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = state_req.special_requirements
        ws.merge_cells(f'A{row}:P{row}')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

@bp.get('/states/<state_code>')
@jwt_required()
def generate_state_excel(state_code):
    """Generate Excel sheet for specific state"""
    state_code = state_code.upper()
    
    # Get state requirements
    state_req = StateRequirement.query.filter_by(state_code=state_code).first()
    
    # Get mobile homes (optionally filter by property if provided)
    property_id = request.args.get('property_id')
    if property_id:
        mobile_homes = MobileHome.query.filter_by(property_id=property_id).all()
    else:
        mobile_homes = MobileHome.query.all()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create the VIN recovery worksheet
    create_vin_recovery_worksheet(wb, state_code, mobile_homes, state_req)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    filename = f"{state_code}_Mobile_Home_VIN_Recovery_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    def remove_file(response):
        try:
            os.unlink(temp_file.name)
        except Exception:
            pass
        return response
    
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.get('/all-states')
@jwt_required()
def generate_all_states_excel():
    """Generate Excel workbook with sheets for all states"""
    
    # Get all state requirements
    all_states = StateRequirement.query.all()
    
    # Get mobile homes (optionally filter by property if provided)
    property_id = request.args.get('property_id')
    if property_id:
        mobile_homes = MobileHome.query.filter_by(property_id=property_id).all()
    else:
        mobile_homes = MobileHome.query.all()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create summary sheet
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws['A1'] = "Mobile Home Park VIN/Title Recovery - All States Summary"
    summary_ws['A1'].font = Font(bold=True, size=16)
    
    summary_ws['A3'] = "Total Mobile Homes:"
    summary_ws['B3'] = len(mobile_homes)
    
    summary_ws['A5'] = "State"
    summary_ws['B5'] = "Total Units"
    summary_ws['C5'] = "Agency"
    summary_ws['D5'] = "Processing Time"
    
    # Summary data
    row = 6
    for state in all_states:
        state_mobile_homes = [mh for mh in mobile_homes if True]  # All for now
        summary_ws[f'A{row}'] = state.state_name
        summary_ws[f'B{row}'] = len(state_mobile_homes)
        summary_ws[f'C{row}'] = state.agency_name or "N/A"
        summary_ws[f'D{row}'] = state.processing_time or "N/A"
        row += 1
    
    # Create individual state sheets
    for state in all_states:
        create_vin_recovery_worksheet(wb, state.state_code, mobile_homes, state)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    filename = f"All_States_Mobile_Home_VIN_Recovery_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.get('/property/<int:property_id>')
@jwt_required()
def generate_property_excel(property_id):
    """Generate Excel sheet for specific property's mobile homes"""
    
    # Verify property exists
    property_obj = Property.query.get_or_404(property_id)
    
    # Get mobile homes for this property
    mobile_homes = MobileHome.query.filter_by(property_id=property_id).all()
    
    if not mobile_homes:
        return jsonify({'error': 'No mobile homes found for this property'}), 404
    
    # Get state from request args or default to first state
    state_code = request.args.get('state_code', 'CA').upper()
    state_req = StateRequirement.query.filter_by(state_code=state_code).first()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create the VIN recovery worksheet
    create_vin_recovery_worksheet(wb, state_code, mobile_homes, state_req)
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    filename = f"Property_{property_id}_{state_code}_VIN_Recovery_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        temp_file.name,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )